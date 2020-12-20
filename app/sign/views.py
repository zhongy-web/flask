import datetime
import math
import os

import openpyxl
from flask import request, current_app, render_template, flash, redirect, url_for
from flask_login import login_required, current_user

from app import db
from app.decorators import permission_required, admin_required
from app.models import Permission, User, Signer
from app.sign import sign
from app.sign.forms import SignForm, ReplenishSignForm, ExcelForm, SignOutForm

'''
    该视图函数进入到用户信息查询页面
'''


@sign.route('/moderate')
@login_required
@permission_required(Permission.MODERATE)
def moderate():
    page = request.args.get('page', 1, type=int)
    pagination = User.query.order_by(User.id.asc()).paginate(
        page, per_page=current_app.config['FLASKY_COMMENTS_PER_PAGE'],
        error_out=False)
    users = pagination.items
    return render_template('sign/cim.html', users=users,
                           pagination=pagination, page=page)


'''
    用户信息页面点击名字进入到个人签到记录页面,
'''


@sign.route('/user_sign/<username>')
@login_required
def user_sign(username):
    user = User.query.filter_by(username=username).first_or_404()
    page = request.args.get('page', 1, type=int)
    pagination = user.signers.order_by(Signer.sign_time.desc()).paginate(
        page, per_page=current_app.config['FLASKY_SIGNERS_PER_PAGE'],
        error_out=False)
    signers = pagination.items
    return render_template('sign/single_sign.html', signers=signers, user=user,
                           pagination=pagination, page=page)


'''
    用于补签的视图函数,本来想加在导航条里，发现不太好拿其他user的数据，
    所有在个人签到页面旁边加了按钮，用于补签,仅管理员可见，且用了保护路由，
    不会被管理员以外权限的用户访问到
'''


@sign.route('/replenish_sign/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def replenish_sign_admin(id):
    user = User.query.get_or_404(id)
    form = ReplenishSignForm(user)
    if form.validate_on_submit():
        user.sign_time_total += form.re_sign.data
        db.session.add(user)
        db.session.commit()
        flash("已为您追加{}小时，下次别忘记签到哦".format(form.re_sign.data))
        return redirect(url_for('sign.user_sign', username=user.username))
    return render_template('sign/replenish_sign.html', form=form, user=user)


'''
这个sign视图用于点击签到按钮后，自动签到用
'''


@sign.route('/sign', methods=['POST'])
def sign_in():
    user = User.query.filter_by(username=current_user.username).first()
    form = SignForm()
    if not user.name:
        flash("请填写您的班级和真实姓名")
        return redirect(url_for('main.edit_profile'))
    if current_user.can(Permission.WRITE) and form.validate_on_submit():
        signer_class = user.s_class
        name = user.name
        signer = Signer(signer_class=signer_class, name=name, signer=user)
        user.sign_status = True
        db.session.add(signer, user)
        db.session.commit()
        flash('%s,你已经签到成功！请勿重复签到！！' % name)
        return redirect(url_for('main.index'))
    else:
        flash('签到失败！！！')
        return redirect(url_for('main.index'))


'''
    新添加一个签退视图，用于实现签退功能
'''


@sign.route('/signout', methods=['POST'])
def sign_out():
    user = User.query.filter_by(username=current_user.username).first()
    sign = Signer.query.filter_by(signer_id=user.id).order_by(Signer.sign_time.desc()).first()
    form = SignOutForm()
    if current_user.can(Permission.WRITE) and form.validate_on_submit():
        user.sign_status = False
        sign.signout_time = datetime.datetime.now()
        t1 = sign.sign_time.strftime("%Y-%m-%d %H:%M:%S")
        t1 = datetime.datetime.strptime(t1, r"%Y-%m-%d %H:%M:%S")
        t2 = sign.signout_time.strftime("%Y-%m-%d %H:%M:%S")
        t2 = datetime.datetime.strptime(t2, r"%Y-%m-%d %H:%M:%S")
        # 时间类型换算问题，使用seconds使之变为秒数，然后换算
        time_total = (t2 - t1).seconds/3600
        # float类型小数点处理
        # user.sign_time_total += round(time_total, 3)
        user.sign_time_total += math.floor(time_total * 10 ** 2) / (10 ** 2)
        db.session.add(user)
        db.session.commit()
        flash('你已经签退成功！请勿重复操作！！')
        return redirect(url_for('main.index'))
    else:
        flash('签退失败！！！')
        return redirect(url_for('main.index'))


'''
    用于生成excel表格
'''


@sign.route("/excel", methods=['POST', 'GET'])
@login_required
@admin_required
def excel():
    form = ExcelForm()
    users = User.query.all()
    if form.validate_on_submit():
        data = datetime.datetime.now().strftime('%Y-%m-%d')  # 获取今日日期
        path = os.path.join(os.getcwd() + "\\" + data + ".xlsx")  # 此路径获取当前目录(flask)下的文件  windows将 '/' 改成 '\\'
        p1 = os.path.exists(path)  # 判断文件是否存在
        wb = openpyxl.Workbook()  # 此处可以赋予读写权限：write_only=True，无法读写
        if not p1:
            wb.save('{}.xlsx'.format(data))
        wb = openpyxl.open('{}.xlsx'.format(data))  # 打开
        # 第一页sheet操作
        cs = wb.create_sheet("全体学习时长记录", 0)  # 页名、页位置（如果页名重复了，会被自动改名）
        wb._active_sheet_index = 0  # 设置活动sheet为哪个一页
        ws = wb.active
        first_row = ['序号', '班级', '姓名', '学习时长']
        for i, j in zip(range(1, 5), first_row):
            ws.cell(1, i, j)
        for user in users:
            ws.append([user.id, user.s_class, user.name, user.sign_time_total])

        # 第二页sheet操作
        cs = wb.create_sheet("未完成周任务记录", 1)  # 创建sheet
        wb._active_sheet_index = 1
        ws = wb.active
        first_row = ['班级', '姓名', '学习时长']
        for i, j in zip(range(1, 4), first_row):
            ws.cell(1, i, j)

        for user in users:
            if user.sign_time_total < 5:
                ws.append([user.s_class, user.name, user.sign_time_total])
        wb.save('{}.xlsx'.format(data))
        wb.close()
        flash("excel生成成功")
    return render_template('sign/excel.html', form=form)

